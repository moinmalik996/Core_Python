"""
Openpyxl is used to read and write Excel Files

"""

import openpyxl

wb = openpyxl.load_workbook("C:\\Users\\moinm\\Desktop\\My Python\\Core_Python\\Openpyxl\\Data.xlsx")

sheets = wb.sheetnames
print(sheets)


sheet1 = wb['name']

# Method_1
print(sheet1['B2'].value)

# Method_2
print(wb['name']['B2'].value)

# Method_3
print(sheet1.cell(2, 1).value)

"""
Now We will work on sheet2 i.e marks
"""

sheet2 = wb['marks']
print(sheet2.cell(2, 2).value)
print(sheet2.cell(3, 2).value)


# Loops to Read Data from Rows and Columns

sheet2 = wb['marks']
rows = sheet2.max_row
columns = sheet2.max_column

print("*" * 30)

for i in range(1, rows + 1):
    for j in range(1, columns + 1):
        print(sheet2.cell(i, j).value)
        print("")

# Writing to Excel File (Giving the Row and Column)

next_row = rows + 1

sheet2.cell(row=next_row, column=1, value="Ali")
sheet2.cell(row=next_row, column=2, value="45.7")
sheet2.cell(row=next_row, column=3, value="9")

wb.save("C:\\Users\\moinm\\Desktop\\My Python\\Core_Python\\Openpyxl\\Data.xlsx")


